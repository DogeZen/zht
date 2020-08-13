
import xlrd,os
from openpyxl import load_workbook

first_day_work,second_day_work,first_day_problem,second_day_problem,first_day_arrange,sencond_day_arrange="","","","","",""

for file_name in os. listdir("汇总"):
    sheet = xlrd.open_workbook("汇总/"+file_name).sheet_by_index(0)
    print(file_name)
    file_name = file_name.replace(".xlsx","").replace("xls","")
    first_day_work+=file_name+":"+sheet.row(4)[2].value+"\n"
    first_day_problem += file_name + ":"+sheet.row(4)[3].value + "\n"
    first_day_arrange += file_name +":"+ sheet.row(4)[4].value + "\n"

    second_day_work +=file_name+":"+sheet.row(5)[2].value+"\n"
    second_day_problem += file_name + ":" + sheet.row(5)[3].value + "\n"
    sencond_day_arrange += file_name + ":" + sheet.row(5)[4].value + "\n"

wb = load_workbook(filename="模板.xlsx")  # 打开excel文件
ws = wb['Sheet1']
ws["B3"] = first_day_work
ws["C3"] = first_day_problem
ws["D3"] = first_day_arrange
ws["B11"] = second_day_work
ws["C11"] = second_day_problem
ws["D11"] = second_day_work
wb.save("汇总文件.xlsx")

print('运行成功，按任意键退出')
c=input()



